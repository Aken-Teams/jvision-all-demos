from pathlib import Path
from openpyxl import load_workbook

path = Path.home() / "Desktop" / "JVision系統Demo清單.xlsx"
wb = load_workbook(path)
ws = wb.active

url = "https://jvision-education-care.vercel.app"
row_values = [
    None,
    "教育照護管理",
    "Education Care Management",
    "招生 CRM、學務排課、出勤接送、收費財務、電子聯絡簿、人事薪資與幼教照護管理",
    url,
]

existing_row = None
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 5).value == url or ws.cell(row, 2).value == row_values[1]:
      existing_row = row
      break

if existing_row is None:
    target_row = None
    for row in range(2, ws.max_row + 2):
        if all(ws.cell(row, col).value in (None, "") for col in range(1, 6)):
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
else:
    target_row = existing_row

max_no = 0
for row in range(2, ws.max_row + 1):
    value = ws.cell(row, 1).value
    if isinstance(value, int):
        max_no = max(max_no, value)

row_values[0] = ws.cell(target_row, 1).value if existing_row else max_no + 1

for col, value in enumerate(row_values, start=1):
    ws.cell(target_row, col).value = value

wb.save(path)
print(f"updated row {target_row}: {row_values}")
