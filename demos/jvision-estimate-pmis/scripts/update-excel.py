from pathlib import Path
from openpyxl import load_workbook

path = Path.home() / "Desktop" / "JVision系統Demo清單.xlsx"
wb = load_workbook(path)
ws = wb.active

url = "https://jvision-estimate-pmis.vercel.app"
row_values = [
    None,
    "估價與工程管理",
    "Estimate + PMIS Engineering Management",
    "工程估價、報價簽核、轉專案、進度管理、品質安全、圖說送審、工程財務與估驗請款",
    url,
]

existing_row = None
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 5).value == url or ws.cell(row, 2).value == row_values[1]:
      existing_row = row
      break

if existing_row is None:
    target_row = None
    for row in range(2, ws.max_row + 2):
        if all(ws.cell(row, col).value in (None, "") for col in range(1, 6)):
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
else:
    target_row = existing_row

max_no = 0
for row in range(2, ws.max_row + 1):
    value = ws.cell(row, 1).value
    if isinstance(value, int):
        max_no = max(max_no, value)

row_values[0] = ws.cell(target_row, 1).value if existing_row else max_no + 1

for col, value in enumerate(row_values, start=1):
    ws.cell(target_row, col).value = value

wb.save(path)
print(f"updated row {target_row}: {row_values}")
